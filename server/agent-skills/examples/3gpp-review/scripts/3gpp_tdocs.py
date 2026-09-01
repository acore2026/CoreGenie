#!/usr/bin/env python3
"""Linux helper for the 3gpp-review skill."""

from __future__ import annotations

import argparse
import concurrent.futures
import hashlib
import json
import posixpath
import re
import shutil
import subprocess
import tempfile
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from lxml import etree

try:
    from openpyxl import load_workbook
except ImportError:  # Conversion mode does not need the Excel dependency.
    load_workbook = None


DOC_RE = re.compile(r"\b([A-Z]\d-\d{6,8})\b", re.I)
NS = {
    "w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "v": "urn:schemas-microsoft-com:vml",
    "o": "urn:schemas-microsoft-com:office:office",
}
ALLOWED_3GPP_HOSTS = {"www.3gpp.org", "ftp.3gpp.org"}
MAX_DOWNLOAD_BYTES = 100 * 1024 * 1024
MAX_ARCHIVE_ENTRY_BYTES = 100 * 1024 * 1024
MAX_ARCHIVE_TOTAL_BYTES = 250 * 1024 * 1024
MANIFEST_SCHEMA = "3gpp-review-manifest/v1"
COVERAGE_SCHEMA = "3gpp-review-coverage/v1"
CONVERSION_SCHEMA = "3gpp-markdown-conversion/v1"


def cell_value(cell) -> str:
    return "" if cell.value is None else str(cell.value).strip()


def choose_sheet(workbook, requested: str | None):
    if requested:
        if requested not in workbook.sheetnames:
            raise SystemExit(f"Sheet not found: {requested!r}; choices={workbook.sheetnames}")
        return workbook[requested]
    if len(workbook.sheetnames) != 1:
        raise SystemExit(f"Specify --sheet; choices={workbook.sheetnames}")
    return workbook[workbook.sheetnames[0]]


def cmd_inspect(args) -> None:
    if load_workbook is None:
        raise SystemExit("openpyxl is required for meeting Index commands")
    workbook = load_workbook(args.excel, read_only=True, data_only=True)
    print("Sheets:")
    for sheet in workbook:
        print(f"  {sheet.title}: rows={sheet.max_row}, columns={sheet.max_column}")
    sheets = [choose_sheet(workbook, args.sheet)] if args.sheet else list(workbook)
    normalized_query = (
        re.sub(r"[^a-z0-9]+", "", args.query.casefold()) if args.query else ""
    )
    for sheet in sheets:
        print(f"\n=== {sheet.title} ===")
        max_row = sheet.max_row if normalized_query else min(args.rows, sheet.max_row)
        matches = 0
        for row_no, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_row), 1
        ):
            values = [cell_value(cell) for cell in row]
            if normalized_query:
                normalized_row = re.sub(
                    r"[^a-z0-9]+", "", " ".join(values).casefold()
                )
                if normalized_query not in normalized_row:
                    continue
                matches += 1
                if matches > args.rows:
                    break
            display_values = values[:10] if normalized_query else values
            fields = [
                f"C{i}={value}"
                for i, value in enumerate(display_values, 1)
                if value
            ]
            if fields:
                print(f"R{row_no}: " + " | ".join(fields))
        if normalized_query and not matches:
            print(f"No rows matched query: {args.query}")
        elif normalized_query and matches > args.rows:
            print(f"Showing the first {args.rows} matching rows.")


def detect_headers(rows: list[list[str]]) -> dict[str, tuple[int, int]]:
    patterns = {
        "status": ("status", "availability"),
        "agenda": ("agenda item", "agenda", "ai"),
        "document": ("tdoc", "document number", "doc number", "tdoc no"),
        "title": ("title", "subject"),
        "source": ("source", "company"),
    }
    found: dict[str, tuple[int, int]] = {}
    for row_idx, row in enumerate(rows):
        for col_idx, raw in enumerate(row):
            text = re.sub(r"\s+", " ", raw.lower()).strip()
            for field, labels in patterns.items():
                if field not in found and any(
                    text == label or text.startswith(label + " ") for label in labels
                ):
                    found[field] = (row_idx, col_idx)
    return found


def find_doc_column(rows: list[list[str]]) -> int | None:
    scores: dict[int, int] = {}
    for row in rows:
        for idx, text in enumerate(row):
            if DOC_RE.search(text):
                scores[idx] = scores.get(idx, 0) + 1
    return max(scores, key=scores.get) if scores else None


def cmd_filter(args) -> None:
    if load_workbook is None:
        raise SystemExit("openpyxl is required for meeting Index commands")
    workbook = load_workbook(args.excel, read_only=True, data_only=True)
    sheet = choose_sheet(workbook, args.sheet)
    rows = [[cell_value(cell) for cell in row] for row in sheet.iter_rows()]
    headers = detect_headers(rows[: min(30, len(rows))])
    doc_col = headers.get("document", (0, find_doc_column(rows)))[1]
    if doc_col is None:
        raise SystemExit("Could not detect a document-number column; inspect the workbook first")
    agenda_col = headers.get("agenda", (0, None))[1]
    title_col = headers.get("title", (0, None))[1]
    source_col = headers.get("source", (0, None))[1]
    status_col = headers.get("status", (0, None))[1]
    start_row = max((position[0] for position in headers.values()), default=0) + 1

    def at(row, col):
        return row[col].strip() if col is not None and col < len(row) else ""

    requested = {
        match.group(1).upper()
        for value in (args.documents or [])
        if (match := DOC_RE.search(value))
    }
    if args.documents and len(requested) != len(args.documents):
        raise SystemExit("Every --documents value must be a valid TDoc number")
    proposals = []
    seen = set()
    for row in rows[start_row:]:
        match = DOC_RE.search(at(row, doc_col))
        if not match:
            continue
        if args.agenda and args.agenda.casefold() not in at(row, agenda_col).casefold():
            continue
        source_filter = getattr(args, "source", None)
        if source_filter and source_filter.casefold() not in at(row, source_col).casefold():
            continue
        doc = match.group(1).upper()
        if requested and doc not in requested:
            continue
        if doc in seen:
            continue
        seen.add(doc)
        proposals.append(
            {
                "document": doc,
                "agenda": at(row, agenda_col),
                "title": at(row, title_col),
                "source": at(row, source_col),
                "status": at(row, status_col),
            }
        )
    missing = sorted(requested - {item["document"] for item in proposals})
    if missing:
        raise SystemExit(
            "Requested TDocs were not found under the selected agenda: "
            + ", ".join(missing)
        )
    payload = {
        "schema": MANIFEST_SCHEMA,
        "excel": str(Path(args.excel).resolve()),
        "sheet": sheet.title,
        "agenda_filter": args.agenda,
        "source_filter": getattr(args, "source", None),
        "count": len(proposals),
        "proposals": proposals,
    }
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Saved {len(proposals)} proposals to {output}")
    for item in proposals:
        print(f"{item['document']}\t{item['source']}\t{item['title']}")


def manifest_payload(path: str) -> dict:
    manifest = Path(path)
    try:
        payload = json.loads(manifest.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise SystemExit(f"Manifest not found: {manifest}") from exc
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Manifest is not valid JSON: {manifest}: {exc}") from exc
    if not isinstance(payload, dict) or payload.get("schema") != MANIFEST_SCHEMA:
        raise SystemExit(
            f"Manifest must be generated by filter-index and use schema {MANIFEST_SCHEMA!r}"
        )
    proposals = payload.get("proposals")
    if not isinstance(proposals, list):
        raise SystemExit("Manifest field 'proposals' must be an array")
    if payload.get("count") != len(proposals):
        raise SystemExit("Manifest field 'count' does not match the proposals array")
    required = ("document", "agenda", "title", "source", "status")
    seen = set()
    for index, item in enumerate(proposals):
        if not isinstance(item, dict):
            raise SystemExit(f"Manifest proposal #{index + 1} must be an object")
        missing = [field for field in required if not isinstance(item.get(field), str)]
        if missing:
            raise SystemExit(
                f"Manifest proposal #{index + 1} has missing or non-string fields: "
                + ", ".join(missing)
            )
        match = DOC_RE.fullmatch(item["document"].strip())
        if not match:
            raise SystemExit(
                f"Manifest proposal #{index + 1} has invalid document: {item['document']!r}"
            )
        document = match.group(1).upper()
        if document in seen:
            raise SystemExit(f"Manifest contains duplicate document: {document}")
        seen.add(document)
        item["document"] = document
    return payload


def load_manifest(path: str) -> list[dict]:
    return manifest_payload(path)["proposals"]


def cmd_validate_manifest(args) -> None:
    payload = manifest_payload(args.manifest)
    print(f"Schema: {payload['schema']}")
    print(f"Proposals: {len(payload['proposals'])}")
    print("Documents: " + ", ".join(item["document"] for item in payload["proposals"]))


def validate_3gpp_url(url: str) -> None:
    parsed = urllib.parse.urlsplit(url)
    if parsed.scheme != "https" or parsed.hostname not in ALLOWED_3GPP_HOSTS:
        raise ValueError("Only HTTPS downloads from official 3GPP hosts are allowed")


def validate_archive(bundle: zipfile.ZipFile) -> None:
    total = 0
    for item in bundle.infolist():
        if item.file_size > MAX_ARCHIVE_ENTRY_BYTES:
            raise ValueError(f"Archive entry exceeds limit: {item.filename}")
        total += item.file_size
        if total > MAX_ARCHIVE_TOTAL_BYTES:
            raise ValueError("Archive uncompressed size exceeds limit")


def download_limited(response, target) -> None:
    validate_3gpp_url(response.geturl())
    declared = response.headers.get("Content-Length")
    if declared and int(declared) > MAX_DOWNLOAD_BYTES:
        raise ValueError("Download exceeds the 100 MiB limit")
    written = 0
    while chunk := response.read(1024 * 1024):
        written += len(chunk)
        if written > MAX_DOWNLOAD_BYTES:
            raise ValueError("Download exceeds the 100 MiB limit")
        target.write(chunk)


def download_one(item: dict, base_url: str, output: Path) -> tuple[str, str]:
    doc = item["document"].upper()
    archive = output / f"{doc}.zip"
    existing = list(output.glob(f"{doc}*.docx"))
    if existing:
        return doc, f"cached {existing[0].name}"
    url = f"{base_url.rstrip('/')}/{doc}.zip"
    try:
        validate_3gpp_url(url)
    except ValueError as exc:
        return doc, f"ERROR {exc}"
    if not archive.exists() or not zipfile.is_zipfile(archive):
        request = urllib.request.Request(url, headers={"User-Agent": "3gpp-review/1.0"})
        try:
            with urllib.request.urlopen(request, timeout=120) as response, archive.open("wb") as target:
                download_limited(response, target)
        except (OSError, ValueError, urllib.error.URLError) as exc:
            archive.unlink(missing_ok=True)
            return doc, f"ERROR download {url}: {exc}"
    if not zipfile.is_zipfile(archive):
        archive.unlink(missing_ok=True)
        return doc, "ERROR invalid ZIP"
    if archive.stat().st_size > MAX_DOWNLOAD_BYTES:
        archive.unlink(missing_ok=True)
        return doc, "ERROR ZIP exceeds download limit"
    with zipfile.ZipFile(archive) as bundle:
        try:
            validate_archive(bundle)
        except ValueError as exc:
            archive.unlink(missing_ok=True)
            return doc, f"ERROR {exc}"
        names = [name for name in bundle.namelist() if name.lower().endswith(".docx")]
        if not names:
            return doc, "ERROR ZIP contains no DOCX"
        for index, name in enumerate(names, 1):
            suffix = "" if len(names) == 1 else f"-{index}"
            destination = output / f"{doc}{suffix}.docx"
            with bundle.open(name) as source, destination.open("wb") as target:
                shutil.copyfileobj(source, target)
    return doc, "downloaded"


def cmd_download(args) -> None:
    output = Path(args.output)
    output.mkdir(parents=True, exist_ok=True)
    proposals = load_manifest(args.manifest)
    failures = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as pool:
        futures = [pool.submit(download_one, item, args.base_url, output) for item in proposals]
        for future in concurrent.futures.as_completed(futures):
            doc, result = future.result()
            print(f"{doc}: {result}")
            failures += result.startswith("ERROR")
    if failures:
        raise SystemExit(f"{failures} download(s) failed")


def xml_text(node) -> str:
    return "".join(node.xpath(".//w:t/text()", namespaces=NS)).strip()


def visio_text(data: bytes) -> list[str]:
    texts = []
    try:
        with zipfile.ZipFile(BytesIO(data)) as visio:
            pages = sorted(
                name for name in visio.namelist()
                if re.search(r"visio/pages/page\d+\.xml$", name)
            )
            for name in pages:
                root = etree.fromstring(visio.read(name))
                for node in root.xpath("//*[local-name()='Text']"):
                    text = "".join(node.itertext()).strip()
                    if text:
                        texts.append(text)
    except (zipfile.BadZipFile, etree.XMLSyntaxError):
        pass
    return texts


def safe_name(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", Path(name).name)


def relationship_map(package: zipfile.ZipFile) -> dict[str, dict[str, str]]:
    try:
        root = etree.fromstring(package.read("word/_rels/document.xml.rels"))
    except KeyError:
        return {}
    relationships = {}
    for node in root:
        rid = node.get("Id")
        if rid:
            relationships[rid] = {
                "target": node.get("Target", ""),
                "type": node.get("Type", ""),
                "mode": node.get("TargetMode", ""),
            }
    return relationships


def package_target(target: str) -> str:
    return posixpath.normpath(posixpath.join("word", target)).lstrip("/")


def markdown_escape(text: str) -> str:
    return text.replace("\\", "\\\\").replace("`", "\\`")


def table_escape(text: str) -> str:
    return markdown_escape(re.sub(r"\s*\n\s*", "<br>", text.strip())).replace("|", "\\|")


def style_names(package: zipfile.ZipFile) -> dict[str, str]:
    try:
        root = etree.fromstring(package.read("word/styles.xml"))
    except KeyError:
        return {}
    names = {}
    for style in root.xpath("./w:style", namespaces=NS):
        style_id = style.get(f"{{{NS['w']}}}styleId", "")
        name = style.find("w:name", NS)
        if style_id:
            names[style_id] = name.get(f"{{{NS['w']}}}val", style_id) if name is not None else style_id
    return names


def unique_asset_name(target: str, used: set[str]) -> str:
    original = safe_name(target) or "image.bin"
    stem, suffix = Path(original).stem, Path(original).suffix
    candidate = original
    index = 2
    while candidate.casefold() in used:
        candidate = f"{stem}-{index}{suffix}"
        index += 1
    used.add(candidate.casefold())
    return candidate


def render_visio(source: Path, assets: Path, stem: str) -> tuple[list[Path], str | None]:
    libreoffice = shutil.which("libreoffice")
    pdftoppm = shutil.which("pdftoppm")
    if not libreoffice or not pdftoppm:
        return [], "缺少 LibreOffice 或 pdftoppm，已保留原始 Visio 文件。"
    try:
        with tempfile.TemporaryDirectory(prefix="3gpp-visio-") as temporary:
            temporary_path = Path(temporary)
            profile = temporary_path / "profile"
            subprocess.run(
                [
                    libreoffice,
                    "--headless",
                    f"-env:UserInstallation={profile.resolve().as_uri()}",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(temporary_path),
                    str(source),
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=90,
            )
            pdf = temporary_path / f"{source.stem}.pdf"
            if not pdf.exists():
                return [], "LibreOffice 没有生成 Visio 预览，已保留原始文件。"
            prefix = temporary_path / "page"
            subprocess.run(
                [pdftoppm, "-png", "-r", "180", str(pdf), str(prefix)],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=90,
            )
            rendered = []
            for index, image in enumerate(sorted(temporary_path.glob("page-*.png")), 1):
                destination = assets / f"{stem}-{index}.png"
                shutil.copyfile(image, destination)
                rendered.append(destination)
            if rendered:
                return rendered, None
            return [], "Visio 预览为空，已保留原始文件。"
    except (OSError, subprocess.SubprocessError) as exc:
        return [], f"Visio 预览生成失败：{exc}"


def convert_docx_to_markdown_legacy(docx: Path, output: Path) -> dict:
    if docx.suffix.lower() != ".docx":
        raise ValueError("Input must be a DOCX file")
    output.mkdir(parents=True, exist_ok=True)
    assets = output / "assets"
    embedded = output / "embedded"
    assets.mkdir(exist_ok=True)
    embedded.mkdir(exist_ok=True)
    warnings: list[str] = []
    image_files: list[str] = []
    embedded_files: list[str] = []
    rendered_visio: list[str] = []
    used_assets: set[str] = set()

    with zipfile.ZipFile(docx) as package:
        validate_archive(package)
        if "word/document.xml" not in package.namelist():
            raise ValueError("DOCX does not contain word/document.xml")
        document = etree.fromstring(package.read("word/document.xml"))
        relationships = relationship_map(package)
        styles = style_names(package)
        extracted_by_rid: dict[str, str] = {}

        def extract_image(rid: str) -> str | None:
            if rid in extracted_by_rid:
                return extracted_by_rid[rid]
            relation = relationships.get(rid)
            if not relation or relation["mode"].lower() == "external":
                return None
            target = package_target(relation["target"])
            if target not in package.namelist():
                warnings.append(f"图片引用 {rid} 指向的文件不存在。")
                return None
            name = unique_asset_name(target, used_assets)
            destination = assets / name
            destination.write_bytes(package.read(target))
            relative = f"assets/{name}"
            extracted_by_rid[rid] = relative
            image_files.append(relative)
            if destination.suffix.lower() in {".emf", ".wmf"}:
                warnings.append(f"{name} 是 {destination.suffix.upper()[1:]} 图片，部分 Markdown 阅读器可能无法显示。")
            return relative

        def inline_content(node) -> str:
            kind = etree.QName(node).localname
            if kind in {"del", "commentRangeStart", "commentRangeEnd"}:
                return ""
            if kind == "t":
                return markdown_escape(node.text or "")
            if kind == "tab":
                return "    "
            if kind in {"br", "cr"}:
                return "  \n"
            if kind == "r":
                content = "".join(inline_content(child) for child in node if etree.QName(child).localname != "rPr")
                properties = node.find("w:rPr", NS)
                if not content:
                    return ""
                if properties is not None and properties.find("w:b", NS) is not None:
                    content = f"**{content}**"
                if properties is not None and properties.find("w:i", NS) is not None:
                    content = f"*{content}*"
                return content
            if kind == "hyperlink":
                content = "".join(inline_content(child) for child in node)
                rid = node.get(f"{{{NS['r']}}}id")
                relation = relationships.get(rid or "")
                if content and relation and relation["mode"].lower() == "external":
                    return f"[{content}]({relation['target']})"
                return content
            if kind in {"drawing", "pict"}:
                references = node.xpath(".//a:blip/@r:embed | .//v:imagedata/@r:id", namespaces=NS)
                values = []
                for rid in references:
                    relative = extract_image(rid)
                    if relative:
                        values.append(f"![文档图片]({relative})")
                return "\n\n".join(values)
            return "".join(inline_content(child) for child in node)

        def paragraph_markdown(paragraph) -> str:
            content = "".join(inline_content(child) for child in paragraph).strip()
            if not content:
                return ""
            properties = paragraph.find("w:pPr", NS)
            style_id = ""
            if properties is not None:
                style = properties.find("w:pStyle", NS)
                if style is not None:
                    style_id = style.get(f"{{{NS['w']}}}val", "")
            style_name = styles.get(style_id, style_id).casefold().replace(" ", "")
            heading = re.search(r"(?:heading|标题)([1-6])", style_name)
            if heading:
                return f"{'#' * int(heading.group(1))} {content}"
            if properties is not None and properties.find("w:numPr", NS) is not None:
                level_node = properties.find("w:numPr/w:ilvl", NS)
                level = int(level_node.get(f"{{{NS['w']}}}val", "0")) if level_node is not None else 0
                return f"{'  ' * level}- {content}"
            return content

        blocks: list[str] = []
        body = document.find("w:body", NS)
        if body is not None:
            for child in body:
                kind = etree.QName(child).localname
                if kind == "p":
                    rendered = paragraph_markdown(child)
                    if rendered:
                        blocks.append(rendered)
                elif kind == "tbl":
                    rows = []
                    for row in child.xpath("./w:tr", namespaces=NS):
                        cells = []
                        for cell in row.xpath("./w:tc", namespaces=NS):
                            cell_parts = [paragraph_markdown(p) for p in cell.xpath("./w:p", namespaces=NS)]
                            cells.append(table_escape("\n".join(part for part in cell_parts if part)))
                        rows.append(cells)
                    if rows:
                        width = max(len(row) for row in rows)
                        normalized = [row + [""] * (width - len(row)) for row in rows]
                        blocks.append("\n".join([
                            "| " + " | ".join(normalized[0]) + " |",
                            "| " + " | ".join(["---"] * width) + " |",
                            *("| " + " | ".join(row) + " |" for row in normalized[1:]),
                        ]))

        used_embedded: set[str] = set()
        for rid, relation in relationships.items():
            relation_type = relation["type"].lower()
            target = package_target(relation["target"])
            if "oleobject" not in relation_type and not target.startswith("word/embeddings/"):
                continue
            if target not in package.namelist():
                warnings.append(f"嵌入对象 {rid} 指向的文件不存在。")
                continue
            name = unique_asset_name(target, used_embedded)
            destination = embedded / name
            destination.write_bytes(package.read(target))
            relative = f"embedded/{name}"
            embedded_files.append(relative)
            if destination.suffix.lower() in {".vsd", ".vsdx"}:
                images, warning = render_visio(destination, assets, f"{destination.stem}-visio")
                for image in images:
                    path_value = f"assets/{image.name}"
                    image_files.append(path_value)
                    rendered_visio.append(path_value)
                if warning:
                    warnings.append(f"{name}：{warning}")
            else:
                warnings.append(f"{name} 无法直接转换，已保留原始嵌入对象。")

    if rendered_visio:
        blocks.append("## 单独导出的 Visio 图\n\n" + "\n\n".join(
            f"![Visio 图 {index}]({path_value})" for index, path_value in enumerate(rendered_visio, 1)
        ))
    markdown_name = f"{safe_name(docx.stem) or 'proposal'}.md"
    markdown_path = output / markdown_name
    markdown_path.write_text("\n\n".join(blocks).strip() + "\n", encoding="utf-8")
    archive_path = output.with_suffix(".zip")
    summary = {
        "schema": CONVERSION_SCHEMA,
        "input": str(docx.resolve()),
        "markdown": markdown_name,
        "images": image_files,
        "embedded": embedded_files,
        "warnings": warnings,
        "archive": str(archive_path.resolve()),
        "convertedAt": datetime.now(timezone.utc).isoformat(),
    }
    summary_path = output / "conversion-summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for item in sorted(output.rglob("*")):
            if item.is_file():
                archive.write(item, item.relative_to(output))
    return summary


def pandoc_docx_to_markdown(docx: Path, output: Path) -> tuple[str, list[str]]:
    pandoc = shutil.which("pandoc")
    if not pandoc:
        raise FileNotFoundError("pandoc is not installed")
    markdown_name = f"{safe_name(docx.stem) or 'proposal'}.md"
    with tempfile.TemporaryDirectory(prefix="3gpp-pandoc-") as temporary:
        temporary_path = Path(temporary)
        execution = subprocess.run(
            [
                pandoc,
                "--from=docx",
                "--to=gfm",
                "--wrap=none",
                "--markdown-headings=atx",
                "--track-changes=accept",
                "--extract-media=assets",
                f"--output={markdown_name}",
                str(docx.resolve()),
            ],
            cwd=temporary_path,
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=180,
        )
        markdown = temporary_path / markdown_name
        if not markdown.is_file() or not markdown.read_text(encoding="utf-8").strip():
            raise ValueError("Pandoc produced an empty Markdown document")
        output.mkdir(parents=True, exist_ok=True)
        shutil.copytree(temporary_path, output, dirs_exist_ok=True)
    warnings = [
        f"Pandoc：{line.strip()}"
        for line in execution.stderr.splitlines()
        if line.strip()
    ]
    return markdown_name, warnings[:20]


def preserve_embedded_objects(
    package: zipfile.ZipFile, output: Path
) -> tuple[list[str], list[str], list[str]]:
    assets = output / "assets"
    embedded = output / "embedded"
    assets.mkdir(exist_ok=True)
    embedded.mkdir(exist_ok=True)
    warnings: list[str] = []
    embedded_files: list[str] = []
    rendered_visio: list[str] = []
    relationships = relationship_map(package)
    targets: list[tuple[str, str]] = []
    seen_targets: set[str] = set()

    for rid, relation in relationships.items():
        relation_type = relation["type"].lower()
        target = package_target(relation["target"])
        if "oleobject" not in relation_type and not target.startswith("word/embeddings/"):
            continue
        if target not in package.namelist():
            warnings.append(f"嵌入对象 {rid} 指向的文件不存在。")
            continue
        if target not in seen_targets:
            seen_targets.add(target)
            targets.append((rid, target))

    for target in package.namelist():
        if (
            target.startswith("word/embeddings/")
            and not target.endswith("/")
            and target not in seen_targets
        ):
            seen_targets.add(target)
            targets.append((target, target))

    used_embedded: set[str] = set()
    for _, target in targets:
        name = unique_asset_name(target, used_embedded)
        destination = embedded / name
        destination.write_bytes(package.read(target))
        embedded_files.append(f"embedded/{name}")
        if destination.suffix.lower() in {".vsd", ".vsdx"}:
            images, warning = render_visio(
                destination, assets, f"{destination.stem}-visio"
            )
            rendered_visio.extend(f"assets/{image.name}" for image in images)
            if warning:
                warnings.append(f"{name}：{warning}")
        else:
            warnings.append(f"{name} 无法直接转换，已保留原始嵌入对象。")
    return embedded_files, rendered_visio, warnings


def output_files(directory: Path, root: Path) -> list[str]:
    if not directory.is_dir():
        return []
    return [
        item.relative_to(root).as_posix()
        for item in sorted(directory.rglob("*"))
        if item.is_file()
    ]


def append_visio_previews(markdown: Path, images: list[str]) -> None:
    if not images:
        return
    content = markdown.read_text(encoding="utf-8").rstrip()
    appendix = "## 单独导出的 Visio 图\n\n" + "\n\n".join(
        f"![Visio 图 {index}]({path_value})"
        for index, path_value in enumerate(images, 1)
    )
    markdown.write_text(f"{content}\n\n{appendix}\n", encoding="utf-8")


def normalize_mixed_ordered_lists(markdown: Path) -> None:
    lines = markdown.read_text(encoding="utf-8").splitlines()
    ordered = re.compile(r"^(?P<indent> {0,12})(?P<number>\d+)[.)][ \t]+")
    escaped = re.compile(
        r"^(?P<indent> {0,12})(?P<number>\d+)\\[.)][ \t]+(?P<content>.*)$"
    )
    normalized: list[str] = []
    previous: tuple[str, int] | None = None
    in_fence = False

    for line in lines:
        stripped = line.lstrip()
        if stripped.startswith("```") or stripped.startswith("~~~"):
            in_fence = not in_fence
            previous = None
            normalized.append(line)
            continue
        if in_fence:
            normalized.append(line)
            continue
        if not line.strip():
            normalized.append(line)
            continue

        manual = escaped.match(line)
        if (
            manual
            and previous
            and manual.group("indent") == previous[0]
            and int(manual.group("number")) == previous[1] + 1
        ):
            while normalized and not normalized[-1].strip():
                normalized.pop()
            number = int(manual.group("number"))
            normalized.append(
                f"{manual.group('indent')}{number}.  {manual.group('content')}"
            )
            previous = (manual.group("indent"), number)
            continue

        item = ordered.match(line)
        previous = (
            (item.group("indent"), int(item.group("number"))) if item else None
        )
        normalized.append(line)

    markdown.write_text("\n".join(normalized).rstrip() + "\n", encoding="utf-8")


def compact_simple_lists(markdown: Path) -> None:
    """Make one-line Pandoc lists tight without joining ordinary paragraphs."""
    lines = markdown.read_text(encoding="utf-8").splitlines()
    manual_bullet = re.compile(
        r"^(?P<prefix> {0,12}(?:> ?)?)(?:\\)(?P<marker>[-+*])[ \t]+"
        r"(?!\\[-+*](?:[ \t]|$))"
    )
    list_item = re.compile(
        r"^(?P<prefix> {0,12}(?:> ?)?)(?P<marker>[-+*]|\d+[.)])[ \t]+"
    )
    normalized: list[str] = []
    in_fence = False

    for line in lines:
        stripped = line.lstrip()
        if stripped.startswith("```") or stripped.startswith("~~~"):
            in_fence = not in_fence
        if not in_fence:
            manual = manual_bullet.match(line)
            if manual:
                line = manual_bullet.sub(
                    f"{manual.group('prefix')}{manual.group('marker')} ",
                    line,
                    count=1,
                )
        normalized.append(line)

    compacted: list[str] = []
    for index, line in enumerate(normalized):
        if not line.strip() and compacted and index + 1 < len(normalized):
            previous = list_item.match(compacted[-1])
            following = list_item.match(normalized[index + 1])
            if previous and following:
                previous_quote = ">" in previous.group("prefix")
                following_quote = ">" in following.group("prefix")
                previous_ordered = previous.group("marker")[0].isdigit()
                following_ordered = following.group("marker")[0].isdigit()
                if (
                    previous_quote == following_quote
                    and previous_ordered == following_ordered
                ):
                    continue
        if (
            line.strip() == ">"
            and compacted
            and index + 1 < len(normalized)
        ):
            previous = list_item.match(compacted[-1])
            following = list_item.match(normalized[index + 1])
            if (
                previous
                and following
                and ">" in previous.group("prefix")
                and ">" in following.group("prefix")
                and previous.group("marker")[0].isdigit()
                == following.group("marker")[0].isdigit()
            ):
                continue
        compacted.append(line)

    markdown.write_text("\n".join(compacted).rstrip() + "\n", encoding="utf-8")


def write_conversion_package(output: Path, summary: dict) -> None:
    summary_path = output / "conversion-summary.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    archive_path = output.with_suffix(".zip")
    with zipfile.ZipFile(archive_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for item in sorted(output.rglob("*")):
            if item.is_file():
                archive.write(item, item.relative_to(output))


def pandoc_fallback_warning(error: Exception) -> str:
    if isinstance(error, FileNotFoundError):
        reason = "运行环境中没有安装 Pandoc"
    elif isinstance(error, subprocess.TimeoutExpired):
        reason = "Pandoc 转换超时"
    elif isinstance(error, subprocess.CalledProcessError):
        reason = f"Pandoc 返回退出码 {error.returncode}"
    else:
        reason = "Pandoc 没有生成有效结果"
    return f"{reason}，已使用兼容转换器。"


def convert_docx_to_markdown(docx: Path, output: Path) -> dict:
    if docx.suffix.lower() != ".docx":
        raise ValueError("Input must be a DOCX file")
    with zipfile.ZipFile(docx) as package:
        validate_archive(package)
        if "word/document.xml" not in package.namelist():
            raise ValueError("DOCX does not contain word/document.xml")

    try:
        markdown_name, warnings = pandoc_docx_to_markdown(docx, output)
    except (OSError, ValueError, subprocess.SubprocessError) as exc:
        summary = convert_docx_to_markdown_legacy(docx, output)
        summary["engine"] = "legacy-ooxml"
        summary["warnings"].insert(0, pandoc_fallback_warning(exc))
        write_conversion_package(output, summary)
        return summary

    with zipfile.ZipFile(docx) as package:
        embedded_files, rendered_visio, embedded_warnings = preserve_embedded_objects(
            package, output
        )
    warnings.extend(embedded_warnings)
    markdown_path = output / markdown_name
    normalize_mixed_ordered_lists(markdown_path)
    compact_simple_lists(markdown_path)
    append_visio_previews(markdown_path, rendered_visio)
    image_files = output_files(output / "assets", output)
    for image in image_files:
        if Path(image).suffix.lower() in {".emf", ".wmf"}:
            warnings.append(
                f"{Path(image).name} 是 {Path(image).suffix.upper()[1:]} 图片，部分 Markdown 阅读器可能无法显示。"
            )

    archive_path = output.with_suffix(".zip")
    summary = {
        "schema": CONVERSION_SCHEMA,
        "engine": "pandoc",
        "input": str(docx.resolve()),
        "markdown": markdown_name,
        "images": image_files,
        "embedded": embedded_files,
        "warnings": warnings,
        "archive": str(archive_path.resolve()),
        "convertedAt": datetime.now(timezone.utc).isoformat(),
    }
    write_conversion_package(output, summary)
    return summary


def cmd_convert_docx(args) -> None:
    try:
        summary = convert_docx_to_markdown(Path(args.input), Path(args.output))
    except (OSError, ValueError, KeyError, zipfile.BadZipFile, etree.XMLSyntaxError) as exc:
        raise SystemExit(f"DOCX conversion failed: {exc}") from exc
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def extract_docx(docx: Path, texts: Path, figures: Path) -> tuple[str, str]:
    doc_id = docx.stem
    doc_figures = figures / doc_id
    doc_figures.mkdir(parents=True, exist_ok=True)
    lines = [f"[SOURCE DOCX: {docx.resolve()}]"]
    try:
        with zipfile.ZipFile(docx) as package:
            validate_archive(package)
            root = etree.fromstring(package.read("word/document.xml"))
            body = root.find("w:body", NS)
            if body is not None:
                for child in body:
                    kind = etree.QName(child).localname
                    if kind == "p":
                        text = xml_text(child)
                        if text:
                            lines.append(text)
                    elif kind == "tbl":
                        lines.append("[TABLE START]")
                        for row in child.xpath("./w:tr", namespaces=NS):
                            cells = [xml_text(cell) for cell in row.xpath("./w:tc", namespaces=NS)]
                            lines.append(" | ".join(cells))
                        lines.append("[TABLE END]")

            xml_root = etree.fromstring(package.read("word/document.xml"))
            shape_text = []
            for node in xml_root.xpath("//*[local-name()='textbox' or local-name()='txbx']"):
                text = xml_text(node)
                if text and text not in shape_text:
                    shape_text.append(text)
            if shape_text:
                lines.extend(["[VML/WPS TEXT]", *shape_text, "[END VML/WPS TEXT]"])

            for name in package.namelist():
                lower = name.lower()
                if lower.startswith("word/media/") and not lower.endswith("/"):
                    destination = doc_figures / safe_name(name)
                    destination.write_bytes(package.read(name))
                    lines.append(f"[FIGURE: {destination.resolve()}]")
                elif lower.startswith("word/embeddings/") and lower.endswith(".vsdx"):
                    data = package.read(name)
                    destination = doc_figures / safe_name(name)
                    destination.write_bytes(data)
                    extracted = visio_text(data)
                    lines.append(f"[VISIO: {destination.resolve()}]")
                    if extracted:
                        lines.extend(["[VISIO TEXT]", *extracted, "[END VISIO TEXT]"])
    except (KeyError, ValueError, zipfile.BadZipFile, etree.XMLSyntaxError) as exc:
        return doc_id, f"ERROR {exc}"
    target = texts / f"{doc_id}.txt"
    target.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return doc_id, f"{target.name}: {len(lines)} lines"


def cmd_extract(args) -> None:
    source = Path(args.input)
    texts = Path(args.texts)
    figures = Path(args.figures)
    texts.mkdir(parents=True, exist_ok=True)
    figures.mkdir(parents=True, exist_ok=True)
    documents = sorted(source.glob("*.docx"))
    if not documents:
        raise SystemExit(f"No DOCX files found in {source}")
    failures = 0
    for docx in documents:
        doc, result = extract_docx(docx, texts, figures)
        print(f"{doc}: {result}")
        failures += result.startswith("ERROR")
    if failures:
        raise SystemExit(f"{failures} extraction(s) failed")


def cmd_coverage(args) -> None:
    manifest = Path(args.manifest)
    receipt = Path(args.receipt)
    payload = manifest_payload(str(manifest))
    expected = {item["document"].upper() for item in payload["proposals"]}
    extracted = set()
    for path in Path(args.texts).glob("*.txt"):
        match = DOC_RE.search(path.stem)
        if match:
            extracted.add(match.group(1).upper())
    missing = sorted(expected - extracted)
    extra = sorted(extracted - expected)
    print(f"Expected: {len(expected)}")
    print(f"Extracted: {len(expected & extracted)}")
    print("Missing: " + (", ".join(missing) if missing else "none"))
    print("Extra: " + (", ".join(extra) if extra else "none"))
    if missing or extra:
        receipt.unlink(missing_ok=True)
        raise SystemExit(1)
    manifest_bytes = manifest.read_bytes()
    coverage = {
        "schema": COVERAGE_SCHEMA,
        "status": "passed",
        "manifest": str(manifest.resolve()),
        "manifestSha256": hashlib.sha256(manifest_bytes).hexdigest(),
        "texts": str(Path(args.texts).resolve()),
        "expectedDocuments": sorted(expected),
        "extractedDocuments": sorted(extracted),
        "missing": [],
        "extra": [],
        "validatedAt": datetime.now(timezone.utc).isoformat(),
    }
    receipt.parent.mkdir(parents=True, exist_ok=True)
    receipt.write_text(
        json.dumps(coverage, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Coverage receipt: {receipt}")


def make_parser() -> argparse.ArgumentParser:
    root = argparse.ArgumentParser(description=__doc__)
    commands = root.add_subparsers(dest="command", required=True)

    inspect = commands.add_parser("inspect-index")
    inspect.add_argument("--excel", required=True)
    inspect.add_argument("--sheet")
    inspect.add_argument("--rows", type=int, default=20)
    inspect.add_argument(
        "--query",
        help="Search all rows for normalized text such as 'KI #18'; --rows limits matching rows",
    )
    inspect.set_defaults(func=cmd_inspect)

    filtering = commands.add_parser("filter-index")
    filtering.add_argument("--excel", required=True)
    filtering.add_argument("--sheet")
    filtering.add_argument("--agenda", required=True)
    filtering.add_argument("--source")
    filtering.add_argument("--documents", nargs="*")
    filtering.add_argument("--output", required=True)
    filtering.set_defaults(func=cmd_filter)

    download = commands.add_parser("download")
    download.add_argument("--manifest", required=True)
    download.add_argument("--base-url", required=True)
    download.add_argument("--output", required=True)
    download.add_argument("--workers", type=int, default=5)
    download.set_defaults(func=cmd_download)

    validating = commands.add_parser("validate-manifest")
    validating.add_argument("--manifest", required=True)
    validating.set_defaults(func=cmd_validate_manifest)

    extract = commands.add_parser("extract")
    extract.add_argument("--input", required=True)
    extract.add_argument("--texts", required=True)
    extract.add_argument("--figures", required=True)
    extract.set_defaults(func=cmd_extract)

    converting = commands.add_parser("convert-docx")
    converting.add_argument("--input", required=True)
    converting.add_argument("--output", required=True)
    converting.set_defaults(func=cmd_convert_docx)

    coverage = commands.add_parser("coverage")
    coverage.add_argument("--manifest", required=True)
    coverage.add_argument("--texts", required=True)
    coverage.add_argument("--receipt", required=True)
    coverage.set_defaults(func=cmd_coverage)
    return root


def main() -> None:
    args = make_parser().parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
